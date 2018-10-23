import re
from openpyxl import load_workbook

s = input('Enter number of column [1]: ')
# if s == '':
#     colum = 1
# else:
#     colum = int(s)
colum = 1 if s == '' else int(s)

# open workbook from file
wb = load_workbook('ss.xlsx')
# grab the active worksheet
ws = wb.active

# go through cells

for i in range(1, ws.max_row+1):
    content = ""
    cc = ws.cell(row=i, column=colum).value
    if type(cc) is str:
        match = re.search(r'[\w\.-]+@[\w\.-]+', cc)
        if match:
            content=match.group(0)
        else:
            content=""
    ws.cell(row=i, column=colum, value=content)

# go through cells
# for i in range(1,4):
#     print('i =', i)
#     for j in range(1,3):
#         print(ws.cell(row=i, column=j).value)

# Save the file
wb.save("ss.xlsx")
print("Done!")
